import numpy as np
import openpyxl
from GPR import GPR
import pyGPs as gp
import matplotlib.pyplot as plt
from data_reader import read_data


filepath = './data_filled.xls'
data, labels = read_data(filepath)

data = data

f = openpyxl.Workbook()
sheet = f.create_sheet()
k = 3

for row_id, d in enumerate(data):
    # if row_id == 72:
    if row_id == 77:
        train_X, train_Y = [], []
        for i in range(len(d)):
            train_X.append(d[i][0])
            train_Y.append(d[i][1])
        train_X = np.array(train_X).reshape(-1, 1)
        train_Y = np.array(train_Y).reshape(-1, 1)
        test_X = np.arange(0, 24, 0.1).reshape(-1, 1)
        gpr = gp.GPR()
        m = gp.mean.Zero()
        c1 = gp.cov.RBF()  # Combine RBF with Matern
        c2 = gp.cov.Matern(d=5)
        gpr.setPrior(mean=m, kernel=c1*c2)
        gpr.setData(train_X, train_Y)
        gpr.optimize(train_X, train_Y)
        mu, cov, _, _, _ = gpr.predict(test_X)
        # if row_id == 57 or row_id == 55 or row_id == 77 or row_id == 121 or row_id==150 or row_id==170:

        gpr.plot()
        test_y = mu.ravel()
        row = [round(test_y[0], k), round(test_y[20], k), round(test_y[40], k), round(test_y[60], k),
               round(test_y[80], k), round(test_y[100], k), round(test_y[120], k), round(test_y[140], k),
               round(test_y[160], k),round(test_y[180], k),round(test_y[200], k),round(test_y[220], k),
               round(test_y[239], k)]
        row_positive = []
        for r in row:
            if r < 0:
                row_positive.append(0)
            else:
                row_positive.append(r)
    # for col_id, item in enumerate(row_positive):
    #     sheet.cell(row=row_id + 2, column=col_id + 1).value = item
# f.save('./data_predict.xls')