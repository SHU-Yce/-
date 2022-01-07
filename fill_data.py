import numpy as np
import openpyxl
from GPR import GPR
from data_reader import read_data
import pyGPs as gp

filepath = './data.xls'
data, labels = read_data(filepath)

f = openpyxl.Workbook()
sheet = f.create_sheet()
k = 3
for row_id, d in enumerate(data):
    train_X, train_Y = [], []
    for i in range(len(d)):
        train_X.append(d[i][0])
        train_Y.append(d[i][1])
    train_X = np.array(train_X).reshape(-1, 1)
    train_Y = np.array(train_Y).reshape(-1, 1)
    test_X = np.arange(0, 12, 0.1).reshape(-1, 1)
    try:
        gpr =gp.GPR()
        m = gp.mean.Linear()
        c1 = gp.cov.RBF()       # Combine RBF with Matern
        c2 = gp.cov.Matern(d=5)
        gpr.setPrior(mean=m, kernel=c1*c2)
        gpr.setData(train_X, train_Y)
        gpr.optimize(train_X, train_Y)
        mu, cov, _, _, _ = gpr.predict(test_X)
        if row_id==8 or row_id==55 or row_id==77:
            gpr.plot()
    except:
        gpr = GPR()
        gpr.fit(train_X, train_Y)
        mu, conv = gpr.predict(test_X)

    test_y = mu.ravel()
    row = [round(test_y[0], k), round(test_y[20], k), round(test_y[40], k), round(test_y[60], k),
           round(test_y[80], k), round(test_y[100], k), round(test_y[119], k)]
    row_positive = []
    for r in row:
        if r < 0:
            row_positive.append(0)
        else:
            row_positive.append(r)
    # print(row_id, row_positive)

    for col_id, item in enumerate(row_positive):
        sheet.cell(row=row_id+2, column=col_id+1).value=item

f.save('./data_filled.xls')

# uncertainty = 1.96 * np.sqrt(np.diag(cov))
# plt.figure()
# plt.title("l=%.2f sigma_f=%.2f" % (gpr.params["l"], gpr.params["sigma_f"]))
# plt.fill_between(test_X.ravel(), test_y + uncertainty, test_y - uncertainty, alpha=0.1)
# plt.plot(test_X, test_y, label="predict")
# plt.scatter(train_X, train_Y, label="train", c="red", marker="x")
# plt.legend()
# plt.show()


